import pandas as pd
from pathlib import Path
from sqlalchemy.orm import Session
from backend.models.database import Task, Measure, Department, User, UserDepartment, RoleEnum, ReportRecord, StatusEnum, TaskLeader, TaskDepartment, TaskPartnerDepartment
from backend.services.auth_service import get_password_hash
import re
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def clean_text(text):
    if pd.isna(text):
        return None
    return str(text).strip()

def import_excel_to_db(db: Session, excel_path: str) -> dict:
    """导入Excel数据到数据库，返回导入统计"""
    df = pd.read_excel(excel_path, header=2)
    df.columns = [
        "序号", "重点工作", "主要目标任务", "牵头领导",
        "牵头部门", "配合部门", "完成时间", "年度工作措施",
        "责任人", "具体举措", "本月工作内容", "下月工作计划"
    ]
    df = df.dropna(subset=["序号"])
    df["序号"] = df["序号"].astype(int)

    stats = {"tasks": 0, "measures": 0, "departments": set()}

    current_task = None
    for _, row in df.iterrows():
        seq = int(row["序号"])
        task_name = clean_text(row["重点工作"])
        target = clean_text(row["主要目标任务"])
        leader = clean_text(row["牵头领导"])
        dept_name = clean_text(row["牵头部门"])
        partner_depts = clean_text(row["配合部门"])
        deadline = clean_text(row["完成时间"])
        measure_content = clean_text(row["年度工作措施"])
        person_liable = clean_text(row["责任人"])

        if pd.isna(seq) or pd.isna(task_name):
            continue

        # 获取或创建部门
        dept = db.query(Department).filter(Department.name == dept_name).first()
        if not dept:
            dept = Department(name=dept_name)
            db.add(dept)
            db.flush()
            stats["departments"].add(dept_name)

        # 创建或更新任务
        if current_task is None or current_task.sequence != seq:
            current_task = Task(
                sequence=seq,
                name=task_name,
                target=target or "",
                deadline=deadline or "",
                year=2026
            )
            db.add(current_task)
            db.flush()
            stats["tasks"] += 1

            # 添加牵头领导
            if leader:
                task_leader = TaskLeader(task_id=current_task.id, leader_name=leader)
                db.add(task_leader)

            # 添加牵头部门
            if dept:
                task_dept = TaskDepartment(task_id=current_task.id, department_id=dept.id)
                db.add(task_dept)

            # 添加配合部门
            if partner_depts:
                for pdept_name in partner_depts.split('\n'):
                    pdept_name = pdept_name.strip()
                    if not pdept_name:
                        continue
                    pdept = db.query(Department).filter(Department.name == pdept_name).first()
                    if not pdept:
                        pdept = Department(name=pdept_name)
                        db.add(pdept)
                        db.flush()
                        stats["departments"].add(pdept_name)
                    task_partner = TaskPartnerDepartment(
                        task_id=current_task.id,
                        department_id=pdept.id,
                        department_name=pdept_name
                    )
                    db.add(task_partner)

        # 创建措施
        if measure_content:
            measure = Measure(
                task_id=current_task.id,
                content=measure_content,
                person_liable=person_liable or ""
            )
            db.add(measure)
            stats["measures"] += 1

    db.commit()
    stats["departments"] = len(stats["departments"])
    return stats

def create_default_users(db: Session):
    """创建默认用户用于测试"""
    # 管理者
    if not db.query(User).filter(User.username == "admin").first():
        admin = User(
            username="admin",
            password_hash=get_password_hash("admin123"),
            roles=RoleEnum.ADMIN
        )
        db.add(admin)

    # 填报者
    if not db.query(User).filter(User.username == "filler").first():
        filler = User(
            username="filler",
            password_hash=get_password_hash("filler123"),
            roles=RoleEnum.FILLER
        )
        db.add(filler)

    # 审批者
    if not db.query(User).filter(User.username == "approver").first():
        approver = User(
            username="approver",
            password_hash=get_password_hash("approver123"),
            roles=RoleEnum.APPROVER
        )
        db.add(approver)

    db.commit()

def export_excel_from_db(db: Session, template_path: str, month: str, selected_columns: list = None) -> bytes:
    """从数据库导出Excel，使用模板文件，填充K列和L列

    Args:
        db: 数据库会话
        template_path: 模板文件路径
        month: 月份 (YYYY-MM格式)
        selected_columns: 选中的列列表，如 ['currentContent', 'nextPlan']

    Returns:
        Excel文件的字节数据
    """
    import re
    import zipfile
    import tempfile

    # 默认选中所有列
    if selected_columns is None:
        selected_columns = ['currentContent', 'nextPlan']

    # 获取当月的所有已审批记录（只有APPROVED状态才导出）
    records = db.query(ReportRecord).filter(
        ReportRecord.month == month,
        ReportRecord.status == StatusEnum.APPROVED
    ).all()

    # 构建 task_sequence -> record 映射
    record_map = {}
    for r in records:
        task = db.query(Task).filter(Task.id == r.task_id).first()
        if task:
            record_map[task.sequence] = r

    # 加载模板 - 需要先修复文件中的filter问题
    # 创建临时文件，移除有问题的filter XML
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name

    with zipfile.ZipFile(template_path, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                content = zin.read(item).decode('utf-8')
                # 移除autoFilter和相关元素 (处理命名空间前缀)
                content = re.sub(r'<(?:ns0:)?autoFilter[^>]*>.*?</(?:ns0:)?autoFilter>', '', content, flags=re.DOTALL)
                content = re.sub(r'<(?:ns0:)?filterColumn[^>]*>.*?</(?:ns0:)?filterColumn>', '', content, flags=re.DOTALL)
                content = re.sub(r'<(?:ns0:)?customFilters>.*?</(?:ns0:)?customFilters>', '', content, flags=re.DOTALL)
                content = re.sub(r'<(?:ns0:)?customFilter[^>]*/>', '', content)
                # 移除autoFilter相关的属性
                content = re.sub(r'\s*filter="[^"]*"', '', content)
                content = re.sub(r'\s*showAutoFilter="[^"]*"', '', content)
                zout.writestr(item, content.encode('utf-8'))

    wb = load_workbook(tmp_path)
    ws = wb['Sheet1']

    # 临时文件会在函数结束时自动清理

    # K列是第11列(索引10)，L列是第12列(索引11)
    K_COL = 11  # 1-based in openpyxl
    L_COL = 12

    # 数据从第4行开始
    DATA_START_ROW = 4

    # 列映射
    COL_MAP = {
        'currentContent': K_COL,
        'nextPlan': L_COL
    }

    # 遍历模板行，填充数据
    for row in range(DATA_START_ROW, 161):
        cell_a = ws.cell(row=row, column=1)
        if not cell_a.value:
            continue

        seq_num = cell_a.value
        matched_record = record_map.get(seq_num)

        if matched_record:
            for col_key in selected_columns:
                if col_key in COL_MAP:
                    col = COL_MAP[col_key]
                    cell = ws.cell(row=row, column=col)
                    if col_key == 'currentContent':
                        cell.value = matched_record.current_content or ''
                    elif col_key == 'nextPlan':
                        cell.value = matched_record.next_plan or ''
                    # 设置字体大小12和自动换行
                    cell.font = Font(size=12, name='宋体')
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 隐藏未选中的列
    for col_key, col in COL_MAP.items():
        if col_key not in selected_columns:
            ws.column_dimensions[get_column_letter(col)].hidden = True

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
